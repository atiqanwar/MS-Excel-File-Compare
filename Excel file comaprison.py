#!/usr/bin/env python
# coding: utf-8

# # Microsoft Excel file comparison script

# ### By: Atiq Anwar

# This script is designed to compare two MS Excel file data (row by row, and column by column). This pythin script reads rows and columns of both input files and compare their values. Diffrence in values identified in both rows and columns will be notified to user at the end of the script.
# 
# Dependencies:
# Python version 3.x
# Modules: Openpyxl, sys

# #### Importing modules for excel file and hashing

# In[13]:


import openpyxl
from openpyxl.utils import get_column_letter
import sys
from os import stat


# #### Opening first file

# In[2]:


file = input("Input first filename:")


# In[3]:


f1 = openpyxl.load_workbook(file)
print ("Worksheets in ", file, "are:", f1.sheetnames)
sheet = input ("Enter sheet name you want to compare from:")


# In[4]:


s1 = f1[sheet]
row = s1.max_row
col = s1.max_column
print ("This sheet contains ", row, " rows and ", col, "columns.")


# #### Opening second file

# In[5]:


file = input("Input second filename:")


# In[6]:


f2 = openpyxl.load_workbook(file)
print ("Worksheets in ", file, "are:", f2.sheetnames)
sheet = input ("Enter sheet name you want to compare with:")


# In[7]:


s2 = f2[sheet]
print ("This sheet contains ", s2.max_row, " rows and ", s2.max_column, "columns.")


# #### Comparing worksheet structures

# In[8]:


if row != s2.max_row and col != s2.max_col:
    print ("Rows and columns in both worksheets are not same.")
    sys.exit(0)


# #### Comparing rows

# In[11]:


diffrow = []
for r in range(1, row+1):
    rowval = []
    rowval2 = []
 #   print (s1[r])
    for cell in s1[r]:
        rowval.append(cell.value)
    for cell in s2[r]:
        rowval2.append(cell.value)
    if rowval != rowval2:
        diffrow.append(r)
print ("Diffrent rows are: ", diffrow)


# #### Comparing columns

# In[18]:


diffcol = []
for c in range(1,col+1):
    colval = []
    colval2 = []
    for cell in s1[get_column_letter(c)]:
        colval.append(cell.value)
    for cell in s2[get_column_letter(c)]:
        colval2.append(cell.value)
    if colval != colval2:
        diffcol.append(get_column_letter(c))
print ("Diffrent columns are: ", diffcol)


# #### Printing output

# In[27]:


if diffcol != []:
    print ("Mismatched cell values are:")
    for c in diffcol:
        for r in diffrow:
            print (c.strip(), str(r).strip(), end="\t")
        print ("")
else:
    print ("No mismatched cell values found.")


# In[ ]:




